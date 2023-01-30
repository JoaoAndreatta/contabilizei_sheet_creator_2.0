import openpyxl as xl
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
import datetime
from dateutil.relativedelta import relativedelta

# creating workbook and worksheet
wb = xl.Workbook()
ws = wb.active
ws.title = 'Empréstimo'

# creating cells variables
a1 = ws['A1']
b1 = ws['B1']
c1 = ws['C1']
d1 = ws['D1']
e1 = ws['E1']
g1 = ws['G1']
h1 = ws['H1']
i1 = ws['I1']
j1 = ws['J1']
k1 = ws['K1']
a2 = ws['A2']
b2 = ws['B2']
c2 = ws['C2']
d2 = ws['D2']
e2 = ws['E2']
g2 = ws['G2']
h2 = ws['H2']
i2 = ws['I2']
j2 = ws['J2']
k2 = ws['K2']

# definig columns names
ws['D1'] = 'CURTO PRAZO'
ws['E1'] = 'LONGO PRAZO'
ws['J1'] = 'CURTO PRAZO'
ws['K1'] = 'LONGO PRAZO'

# defining columns widths
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 15
ws.column_dimensions['G'].width = 3
ws.column_dimensions['H'].width = 12
ws.column_dimensions['I'].width = 12
ws.column_dimensions['J'].width = 15
ws.column_dimensions['K'].width = 15

# creating the bold font
font = Font(bold=True)

# painting the cells bg
yellow_fill = PatternFill(fill_type='solid', start_color='00FFFF00')
a1.fill = yellow_fill
b1.fill = yellow_fill
c1.fill = yellow_fill
d1.fill = yellow_fill
e1.fill = yellow_fill
a2.fill = yellow_fill
b2.fill = yellow_fill
c2.fill = yellow_fill
d2.fill = yellow_fill
e2.fill = yellow_fill
g1.fill = yellow_fill
h1.fill = yellow_fill
i1.fill = yellow_fill
j1.fill = yellow_fill
k1.fill = yellow_fill
g2.fill = yellow_fill
h2.fill = yellow_fill
i2.fill = yellow_fill
j2.fill = yellow_fill
k2.fill = yellow_fill

# apllying custom font
d1.font = font
e1.font = font
b2.font = font
c2.font = font
d2.font = font
e2.font = font
j1.font = font
k1.font = font
h2.font = font
i2.font = font
j2.font = font
k2.font = font

# get values manually
date = '13/09/2022'
date = datetime.datetime.strptime(date,'%d/%m/%Y')
installments = 37
value = 50000
value_installments = float(1561.17)

# defining some predefined values
ws['B3'] = date
ws['B2'] = date - relativedelta(years=1)
ws['C2'] = value
ws['D2'] = "=SUM(C3:C25)"
ws['E2'] = f"=SUM(C26:C{installments + 2})"
ws['H2'] = date - relativedelta(years=1)
ws['H3'] = date
ws['I2'] = "=SUM(D2+E2)-C2"
ws['J2'] = "=SUM(I3:I25)"
ws['K2'] = f"=SUM(I26:I{installments + 2})"

# variables declarations
n = 1
term_1 = 2
term_2 = 3
term_3 = 26

# installments number column
for row in range(3, installments + 3):
    cell = ws.cell(row,1)
    cell.value = n
    n += 1

# date column
for row in range(3, ws.max_row + 1):
    cell = ws.cell(row,2)
    cell.value = date.date() 
    # increases by one the month
    date = date + relativedelta(months=1) 

# installments value column
for row in range(3, ws.max_row + 1):
    cell = ws.cell(row,3)
    cell.value = value_installments

# short term column
for row in range(3, ws.max_row + 1):
    cell = ws.cell(row,4)
    cell.value = f'=D{term_1}-C{term_2}+C{term_3}'
    term_1 += 1 
    term_2 += 1 
    term_3 += 1 

# redefining variables
term_1 = 2
term_2 = 3
term_3 = 26

# long term column
for row in range(3, ws.max_row + 1):
    cell = ws.cell(row,5)
    cell.value = f'=E{term_1}-C{term_3}'
    term_1 += 1
    term_3 += 1

# taxes side

# variables declaration
n = 1
term_1 = 2
term_2 = 3
term_3 = 26
date = '13/09/2022'
date = datetime.datetime.strptime(date,'%d/%m/%Y')

# installments number column
for row in range(3, installments + 3):
    cell = ws.cell(row,7)
    cell.value = n
    n += 1

# date column
for row in range(3, ws.max_row + 1):
    cell = ws.cell(row,8)
    cell.value = date.date() 
    # increases by one the month
    date = date + relativedelta(months=1)

# short term column
for row in range(3, ws.max_row + 1):
    cell = ws.cell(row,9)
    cell.value = f'=I2/{installments}'

# installments value column
for row in range(3, ws.max_row + 1):
    cell = ws.cell(row,10)
    cell.value = f'=J{term_1}-I{term_2}+I{term_3}'
    term_1 += 1 
    term_2 += 1 
    term_3 += 1 

# redefining variables
term_1 = 2
term_2 = 3
term_3 = 26

for row in range(3, ws.max_row + 1):
    cell = ws.cell(row,11)
    cell.value = f'=K{term_1}-I{term_3}'
    term_1 += 1
    term_3 += 1

# redefining variables
term_1 = 2
term_3 = 26

wb.save('test.xlsx')