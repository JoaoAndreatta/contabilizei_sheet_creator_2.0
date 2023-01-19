import tkinter
import customtkinter
from tkinter import filedialog
import openpyxl as xl
import datetime
from dateutil.relativedelta import relativedelta
from openpyxl.styles import NamedStyle

class Depreciacao(customtkinter.CTk):

    def depreciacao_button(self):
            
            # font variables
            font_principal = customtkinter.CTkFont(family="Helvetica", size=23, weight="bold")
            
            # sheet entries frame
            self.entries_frame = customtkinter.CTkFrame(self.main_frame, corner_radius=0, fg_color='transparent')
            self.entries_frame.grid(row=0, column=1, rowspan=3 ,padx=(20, 20), pady=(0, 0), sticky="nsew")
            
            # depreciation widgets
            self.depreciacao_label = customtkinter.CTkLabel(self.entries_frame, text='Criar planílha de depreciação', font=font_principal, text_color='#3b8ed0')
            self.depreciacao_label.grid(row=0, column=1, sticky="nw", pady=(20, 0))
            self.date_entry = customtkinter.CTkEntry(self.entries_frame, placeholder_text="Insira a data de inicio", 
                                                    width=200, height=40, border_width=2)
            self.date_entry.grid(row=1, column=1, sticky="nw", pady=(20, 0))
            self.value_entry = customtkinter.CTkEntry(self.entries_frame, placeholder_text="Insira o valor total", 
                                                    width=200, height=40, border_width=2)
            self.value_entry.grid(row=2, column=1, sticky="nw", pady=(20, 0))
            self.deprec_types_option_menu = customtkinter.CTkOptionMenu(self.entries_frame, width=200, height=40,
                                                                        values=['Veículos','Móveis','Equipamentos','Imóveis','Instalações'],
                                                                        dropdown_fg_color='#3b8ed0', dropdown_hover_color='#36719f', 
                                                                        dropdown_text_color='white')
            self.deprec_types_option_menu.set('Selecionar depreciação')
            self.deprec_types_option_menu.grid(row=3, column=1, sticky="nw", pady=(20, 0))

            # create sheet buttons frame
            self.create_buttons_frame = customtkinter.CTkFrame(self.entries_frame, corner_radius=0, fg_color='transparent')
            self.create_buttons_frame.grid(row=4, column=2 ,padx=(20, 0), pady=(35, 0), sticky="se")

            # create sheet widget
            self.confirm_sheet_button = customtkinter.CTkButton(self.create_buttons_frame, text='Confirmar entradas', 
                                                                command= lambda: [Depreciacao.create_sheet_button_function(self),
                                                                Depreciacao.get_entries(self)])
            self.confirm_sheet_button.grid(row=2, column=1, sticky="s", padx=(0, 160), pady=(0,0))
            self.create_sheet_button = customtkinter.CTkButton(self.create_buttons_frame, text='Criar!', state='disabled', fg_color='grey',
                                                                command= lambda: [Depreciacao.set_save_destination(),
                                                                                Depreciacao.create_depreciation_sheet(self),
                                                                                Depreciacao.created_sucess(self)])
            self.create_sheet_button.grid(row=2, column=1, sticky="s", padx=(160, 0), pady=(0, 0))

            # create sheet preview frame
            self.sheet_frame = customtkinter.CTkFrame(self.main_frame, corner_radius=0)
            self.sheet_frame.grid(row=3, column=1, columnspan=2, padx=(20, 20), pady=(20, 20), sticky="nsew")

            self.not_avaiable = customtkinter.CTkLabel(self.sheet_frame, text='Função ainda não disponível!', height=250, font=font_principal, text_color='grey')
            self.not_avaiable.grid(row=3,column=1, padx=(160,0), sticky='nswe')

    # create depreciation sheet function
    def create_depreciation_sheet(self):

        # variables declaration
        num = 1
        value = Depreciacao.get_entries(self)[0]
        installments = Depreciacao.get_entries(self)[1]
        type = Depreciacao.get_entries(self)[2]
        date = Depreciacao.get_entries(self)[3]
        debit_account = '3.01.01.08.01.11'
        credit_account = Depreciacao.get_entries(self)[4]
        description = Depreciacao.get_entries(self)[5]
        monthly_value = round(float(value) / int(installments), 2)
        
        # creating workbook and worksheet
        wb = xl.Workbook()
        ws = wb.active
        ws.title = 'Depreciação'

        # definig columns names
        ws['A1'] = 'DATA'
        ws['B1'] = 'DESCRIÇÃO'
        ws['C1'] = 'VALOR'
        ws['D1'] = 'CONTA DÉBITO'
        ws['E1'] = 'CONTA CRÉDITO'

        # defining columns widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 36
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15

        # formatating date
        final_date = date
        formated_date = datetime.datetime.strptime(final_date,'%d/%m/%Y') 
        
        # defining date column
        for row in range(2, int(installments) + 2):
            cell = ws.cell(row,1)
            cell.value = formated_date.date() 
            # increases by one the month
            formated_date = formated_date + relativedelta(months=1) 
            
            # this if and else makes it always the last day of the month 
            if formated_date.month == 12:
                formated_date = formated_date.replace(day=31)
            else:
                formated_date = formated_date.replace(month=formated_date.month+1, day=1) - datetime.timedelta(days=1)

        # defining description column
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row,2)
            cell.value = description + str(num) + "/" + str(installments) 
            num += 1

        # defining value column
        for row in range(2,ws.max_row + 1):
            cell = ws.cell(row,3)
            cell.value = monthly_value 

        # defining debit account column
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row,4)
            cell.value = debit_account 

        # defining credit account column
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row,5)
            cell.value = credit_account 

        #Criar o arquivo excel
        wb.save(str(save_spot) + '/nova_planilha.xlsx')

    # get entries function
    def get_entries(self):
        
        # variables declarations
        description = ''
        depreciation_list = ['Veículos','Móveis','Equipamentos','Imóveis','Instalações']
        installments = 0
        credit_account = ''
        debit_account = ''

        value = self.value_entry.get()
        date = self.date_entry.get()
        type = self.deprec_types_option_menu.get()
        if type == depreciation_list[0]:
            type = 0.2
            description = 'Depreciação de veículos '
            credit_account = '1.07.04.12.01'
            installments = 60
        elif type == depreciation_list[1]:
            type = 0.1
            description = 'Depreciação de móveis '
            credit_account = '1.07.04.12.02'
            installments = 120
        elif type == depreciation_list[2]:
            type = 0.2
            description = 'Depreciação de equipamentos '
            credit_account = '1.07.04.12.03'
            installments = 60
        elif type == depreciation_list[3]:
            type = 0.04
            description = 'Depreciação de imóveis '
            credit_account = '1.07.04.12.06'
            installments = 240
        elif type == depreciation_list[4]:
            type = 0.1
            description = 'Depreciação de instalações '
            credit_account = '1.07.04.12.04'
            installments = 120
        
        return value, installments, type, date, credit_account, description

    # function to save created sheet
    def set_save_destination():
        global save_spot
        save_spot = filedialog.askdirectory()
        save_spot = str(save_spot)

    # function to enable the create button
    def create_sheet_button_function(self):
        if (len(self.date_entry.get()) and len(self.value_entry.get()) != 0) and self.deprec_types_option_menu.get() in [
                                                                    'Veículos','Móveis','Equipamentos','Imóveis','Instalações']:
            self.create_sheet_button.configure(fg_color='#5cb85c', hover_color='#4D994D', state='normal')
        else:
            self.create_sheet_button.configure(fg_color='grey')
