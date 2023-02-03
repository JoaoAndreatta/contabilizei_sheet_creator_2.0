import tkinter
import customtkinter
import openpyxl as xl
from tkinter import filedialog
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
import datetime
from dateutil.relativedelta import relativedelta


class Emprestimo():

    def loan_button(self):
        
        # font variables
        font_principal = customtkinter.CTkFont(family="Helvetica", size=23, weight="bold")

        # sheet entries frame
        self.entries_frame = customtkinter.CTkFrame(self.main_frame, corner_radius=0, fg_color='transparent')
        self.entries_frame.grid(row=1, column=1, rowspan=3 ,padx=(20, 20), pady=(20, 0), sticky="nsew")

        # loan widgets
        self.depreciacao_label = customtkinter.CTkLabel(self.main_frame, text='Criar planilha de empréstimo', font=font_principal, text_color='#3b8ed0')
        self.depreciacao_label.grid(row=0, column=1, sticky="nw", padx=(20, 0), pady=(20, 0))
        self.date_entry = customtkinter.CTkEntry(self.entries_frame, placeholder_text="Insira a data de retirada", 
                                                width=200, height=40, border_width=2)
        self.date_entry.grid(row=1, column=1, sticky="nw", pady=(0, 0))
        self.pay_date_entry = customtkinter.CTkEntry(self.entries_frame, placeholder_text="Insira a data da 1ª parcela", 
                                                width=200, height=40, border_width=2)
        self.pay_date_entry.grid(row=2, column=1, sticky="nw", pady=(20, 0))
        self.installments_entry = customtkinter.CTkEntry(self.entries_frame, placeholder_text="Insira o número de parcelas", 
                                                width=200, height=40, border_width=2)
        self.installments_entry.grid(row=3, column=1, sticky="nw", pady=(20, 0))
        self.value_entry = customtkinter.CTkEntry(self.entries_frame, placeholder_text="Insira o valor líquido",
                                                width=200, height=40, border_width=2)
        self.value_entry.grid(row=1, column=2, sticky="nw", padx=(20, 0), pady=(0, 0))
        self.installments_value_entry = customtkinter.CTkEntry(self.entries_frame, placeholder_text="Insira o valor da parcela",
                                                width=200, height=40, border_width=2)
        self.installments_value_entry.grid(row=2, column=2, sticky="nw", padx=(20, 0), pady=(20, 0))

        # create sheet buttons frame
        self.create_buttons_frame = customtkinter.CTkFrame(self.main_frame, corner_radius=0, fg_color='transparent')
        self.create_buttons_frame.grid(row=4, column=1, padx=(365, 0), pady=(20, 0), sticky="se")

        # create sheet widget
        self.confirm_sheet_button = customtkinter.CTkButton(self.create_buttons_frame, text='Confirmar entradas',
                                                        command= lambda: [Emprestimo.create_sheet_button_function(self),
                                                        Emprestimo.get_entries(self)])
        self.confirm_sheet_button.grid(row=0, column=0, padx=(0, 0), pady=(0,0))
        self.create_sheet_button = customtkinter.CTkButton(self.create_buttons_frame, text='Criar!', state='disabled', fg_color='grey',
                                                        command= lambda: [Emprestimo.set_save_destination(),
                                                        Emprestimo.create_loan_sheet(self)])
        self.create_sheet_button.grid(row=0, column=1, padx=(20, 0), pady=(0, 0))

    # create depreciation sheet function
    def create_loan_sheet(self):

        # creating workbook and worksheet
        wb = xl.Workbook()
        ws = wb.active
        ws.title = 'Empréstimo'

        # creating cells variables
        a1 = ws['A1']
        b1 = ws['B1']
        c1 = ws['C1']
        d1 = ws['D1']
        e1 = ws['E1']
        g1 = ws['G1']
        h1 = ws['H1']
        i1 = ws['I1']
        j1 = ws['J1']
        k1 = ws['K1']
        a2 = ws['A2']
        b2 = ws['B2']
        c2 = ws['C2']
        d2 = ws['D2']
        e2 = ws['E2']
        g2 = ws['G2']
        h2 = ws['H2']
        i2 = ws['I2']
        j2 = ws['J2']
        k2 = ws['K2']

        # definig columns names
        ws['D1'] = 'CURTO PRAZO'
        ws['E1'] = 'LONGO PRAZO'
        ws['J1'] = 'CURTO PRAZO'
        ws['K1'] = 'LONGO PRAZO'

        # defining columns widths
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['G'].width = 3
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 15

        # creating the bold font
        font = Font(bold=True)

        # painting the cells bg
        yellow_fill = PatternFill(fill_type='solid', start_color='00FFFF00')
        a1.fill = yellow_fill
        b1.fill = yellow_fill
        c1.fill = yellow_fill
        d1.fill = yellow_fill
        e1.fill = yellow_fill
        a2.fill = yellow_fill
        b2.fill = yellow_fill
        c2.fill = yellow_fill
        d2.fill = yellow_fill
        e2.fill = yellow_fill
        g1.fill = yellow_fill
        h1.fill = yellow_fill
        i1.fill = yellow_fill
        j1.fill = yellow_fill
        k1.fill = yellow_fill
        g2.fill = yellow_fill
        h2.fill = yellow_fill
        i2.fill = yellow_fill
        j2.fill = yellow_fill
        k2.fill = yellow_fill

        # apllying custom font
        d1.font = font
        e1.font = font
        b2.font = font
        c2.font = font
        d2.font = font
        e2.font = font
        j1.font = font
        k1.font = font
        h2.font = font
        i2.font = font
        j2.font = font
        k2.font = font

        # get values manually
        initial_date = Emprestimo.get_entries(self)[2]
        initial_date = datetime.datetime.strptime(initial_date,'%d/%m/%Y')
        pay_date = Emprestimo.get_entries(self)[3]
        pay_date = datetime.datetime.strptime(pay_date,'%d/%m/%Y')
        installments = Emprestimo.get_entries(self)[4]
        installments = int(installments)
        value = Emprestimo.get_entries(self)[0]
        value_installments = Emprestimo.get_entries(self)[1]
        value_installments = value_installments.replace(',','.')
        value_installments = float(value_installments)

        # to see until what row will be the short term
        def diff_month(d1, d2):
            return (d1.year - d2.year) * 12 + d1.month - d2.month
        
        if initial_date == pay_date:
            # 24 months of short term
            start_term = 24
        else:
            # less than 24 months
            start_term = diff_month(pay_date, initial_date)

        # defining some predefined values and styles
        ws['B3'] = pay_date
        ws['B3'].number_format = 'dd-mm-yyyy'
        ws['B2'] = initial_date
        ws['B2'].number_format = 'dd-mm-yyyy'
        ws['C2'] = value
        ws['C2'].number_format = '#,##0.00'
        ws['D2'] = f"=SUM(C3:C{start_term + 2})"
        ws['E2'] = f"=SUM(C{start_term + 3}:C{installments + 2})"
        ws['H2'] = initial_date
        ws['H2'].number_format = 'dd-mm-yyyy'
        ws['H3'] = pay_date
        ws['H3'].number_format = 'dd-mm-yyyy'
        ws['I2'] = "=SUM(D2+E2)-C2"
        ws['J2'] = f"=SUM(I3:I{start_term + 2})"
        ws['K2'] = f"=SUM(I{start_term + 3}:I{installments + 2})"

        # variables declarations
        n = 1
        term_1 = 2
        term_2 = 3
        term_3 = start_term + 2

        # installments number column
        for row in range(3, installments + 3):
            cell = ws.cell(row,1)
            cell.value = n
            n += 1

        # date column
        for row in range(3, ws.max_row + 1):
            cell = ws.cell(row,2)
            cell.value = pay_date.date()
            cell.number_format = 'dd-mm-yyyy' 
            # increases by one the month
            pay_date += relativedelta(months=1) 

        # installments value column
        for row in range(3, ws.max_row + 1):
            cell = ws.cell(row,3)
            cell.value = value_installments
            cell.number_format = '#,##0.00'

        # short term column
        for row in range(3, ws.max_row + 1):
            cell = ws.cell(row,4)
            cell.value = f'=D{term_1}-C{term_2}+C{term_3 + 1}'
            cell.number_format = '#,##0.00'
            term_1 += 1 
            term_2 += 1 
            term_3 += 1 

        # redefining variables
        term_1 = 2
        term_2 = 3
        term_3 = start_term + 2

        # long term column
        for row in range(3, ws.max_row + 1):
            cell = ws.cell(row,5)
            cell.value = f'=E{term_1}-C{term_3 + 1}'
            cell.number_format = '#,##0.00'
            term_1 += 1
            term_3 += 1

        # taxes side

        # variables declaration
        n = 1
        term_1 = 2
        term_2 = 3
        term_3 = start_term + 2
        pay_date = Emprestimo.get_entries(self)[3]
        pay_date = datetime.datetime.strptime(pay_date,'%d/%m/%Y')

        # installments number column
        for row in range(3, installments + 3):
            cell = ws.cell(row,7)
            cell.value = n
            n += 1

        # date column
        for row in range(3, ws.max_row + 1):
            cell = ws.cell(row,8)
            cell.value = pay_date.date()
            cell.number_format = 'dd-mm-yyyy' 
            # increases by one the month
            pay_date += relativedelta(months=1)

        # short term column
        for row in range(3, ws.max_row + 1):
            cell = ws.cell(row,9)
            cell.value = f'=I2/{installments}'
            cell.number_format = '#,##0.00'

        # installments value column
        for row in range(3, ws.max_row + 1):
            cell = ws.cell(row,10)
            cell.value = f'=J{term_1}-I{term_2}+I{term_3 + 1}'
            cell.number_format = '#,##0.00'
            term_1 += 1 
            term_2 += 1 
            term_3 += 1 

        # redefining variables
        term_1 = 2
        term_2 = 3
        term_3 = start_term + 2

        for row in range(3, ws.max_row + 1):
            cell = ws.cell(row,11)
            cell.value = f'=K{term_1}-I{term_3 + 1}'
            cell.number_format = '#,##0.00'
            term_1 += 1
            term_3 += 1

        # redefining variables
        term_1 = 2
        term_3 = start_term + 2

        wb.save(str(save_spot) + '/nova_planilha_emprestimo.xlsx')
        
    
    def get_entries(self):
        
        #get values
        value = self.value_entry.get()
        installments_value = self.installments_value_entry.get()
        initial_date = self.date_entry.get()
        pay_date = self.pay_date_entry.get()
        installments = self.installments_entry.get()

        return value, installments_value, initial_date, pay_date, installments

    # function to save created sheet
    def set_save_destination():
        global save_spot
        save_spot = filedialog.askdirectory()
        save_spot = str(save_spot)
    
    # function to enable the create button
    def create_sheet_button_function(self):
        if (len(self.value_entry.get()) and len(self.installments_value_entry.get())
        and len(self.date_entry.get()) and len(self.pay_date_entry.get()) and len(self.installments_entry.get()) != 0):
            self.create_sheet_button.configure(fg_color='#5cb85c', hover_color='#4D994D', state='normal')
        else:
            self.create_sheet_button.configure(fg_color='grey')